import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter


file_path = 'your_file_path_here.xlsx' 
df = pd.read_excel(file_path, engine='openpyxl')


def fill_missing_values(df):
    for column in df.columns:
        if df[column].isnull().any():  
            if df[column].dtype == 'float64' or df[column].dtype == 'int64':
                df[column].fillna(df[column].mean(), inplace=True)
            elif df[column].dtype == 'object':
                df[column].fillna('unknown', inplace=True)
    return df

def convert_to_date_time(df):
    for column in df.columns:
        if "date" in column.lower():
            df[column] = pd.to_datetime(df[column], errors='coerce')
    return df

def clean_headers(df):
    df.columns = df.columns.str.replace(r'[^\w\s]', '', regex=True).str.strip()
    df.columns = df.columns.str.lower().str.replace(' ', '_')
    return df

df = fill_missing_values(df)          
df = convert_to_date_time(df)        
df = clean_headers(df)                

df.drop_duplicates(inplace=True)

missing_data = df.isnull().sum()
print("\nMissing Data:\n", missing_data)

print(f"\nNumber of duplicate rows: {df.duplicated().sum()}")

cleaned_file_path = 'cleaned_data.xlsx'
df.to_excel(cleaned_file_path, index=False, engine='openpyxl')
print(f"Data cleaning complete. Cleaned data saved to '{cleaned_file_path}'.")

for col in range(1, sheet.max_column + 1):
    col_letter = get_column_letter(col)
    sheet.column_dimensions[col_letter].width = 20
    header_cell = sheet[f'{col_letter}1']
    header_cell.alignment = Alignment(horizontal='center', vertical='center')

from openpyxl.styles import Border, Side

no_border = Border(left=Side(border_style=None),
                   right=Side(border_style=None),
                   top=Side(border_style=None),
                   bottom=Side(border_style=None))

for row in sheet.iter_rows():
    for cell in row:
        cell.border = no_border


sheet._images = []


workbook.save(cleaned_file_path)
print(f"Data cleaning and formatting complete. Cleaned file saved as '{cleaned_file_path}'.")
